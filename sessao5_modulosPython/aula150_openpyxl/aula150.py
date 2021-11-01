"""
https://openpyxl.readthedocs.io/en/stable/
pip install openpyxl
pipenv instal openpyxl
"""

import openpyxl
from random import uniform

"""
pedidos = openpyxl.load_workbook('excel/pedidos.xlsx')
nome_planilha = pedidos.sheetnames
# print(nome_planilha)
planilha1 = pedidos['Página1']

# print(planilha1['b4'].value)


# for campo in planilha1['b']:
#     print(campo.value)

#
# for linha in planilha1['a1:c2']:
#     for coluna in linha:
#         print(coluna.value)
# #
# for linha in planilha1:
#     if linha[0].value is not None:
#         print(linha[0].value, end='\t\t')
#     if linha[1].value is not None:
#         print(linha[1].value, end='\t\t')
#     if linha[2].value is not None:
#         print(linha[2].value)
#     # if linha[3].value is not None:
#     #     print(linha[3].value)
#
#         # print(linha[0].value, linha[1].value, linha[2].value, linha[3].value, sep='\t\t')


# planilha1['b3'].value = 2200

for linha in range(5, 16):
    numero_pedido = linha - 1
    planilha1.cell(linha, 1).value = numero_pedido
    planilha1.cell(linha, 2).value = 1200 + linha
    # print(linha)

    preco = round(uniform(10, 100), 2)
    planilha1.cell(linha, 3).value = preco

pedidos.save('nova_planilha.xlsx')
"""

planilha = openpyxl.Workbook()
planilha.create_sheet('planilha1', 0)
planilha.create_sheet('planilha2', 1)

planilha1 = planilha['planilha1']
planilha2 = planilha['planilha2']

for linha in range(1, 11):
    numero_pedido = linha - 1
    planilha1.cell(linha, 1).value = numero_pedido
    planilha1.cell(linha, 2).value = 1200 + linha

    preco = round(uniform(10, 100), 2)
    planilha1.cell(linha, 3).value = preco

for linha in range(1, 11):
    planilha2.cell(linha, 1).value = f'Luiz {linha} {round(uniform(10, 100),2)}'
    planilha2.cell(linha, 2).value = f'otvio {linha} {round(uniform(10, 100),2)}'
    planilha2.cell(linha, 3).value = f'joaozinho {linha} {round(uniform(10, 100),2)}'


planilha.save('nova_planilha2.xlsx')